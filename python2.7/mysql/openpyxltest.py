#!/usr/bin/env python
# -*- coding: utf-8 -*-


#sudo pip install openpyxl

from openpyxl import Workbook

def main():
  wb = Workbook()
  ws = wb.active
  ws2 = wb.create_sheet()
  ws2['A1']="111,333,444,555,666"
  ws['B4'] = 4444
  # ws1 = wb.create_sheet()
  # ws2 = wb.create_sheet(1)
  ws.title = "Python"
  ws01 = wb['Python']
  print(ws is ws01)
  # ws02 = wb.get_sheet_by_name("Python") # 这个方法名字也太直接了，方法的参数就是 sheet 名字
  # print(ws is ws02)
  # ws01 = wb['Python']    #sheet 和工作簿的关系，类似键值对的关系
  wb.save("mysql/23401.xlsx")
  for sh in wb:
    print sh.title

if __name__ == '__main__':
  main()
